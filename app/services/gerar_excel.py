import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

def gerar_relatorio_excel(json_path="registros_presenca.json", output="relatorio_presenca.xlsx"):
    df = pd.read_json(json_path)
    df.sort_values(by=["data", "hora"], inplace=True)

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Presenças')

        wb = writer.book
        ws = writer.sheets['Presenças']

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")

        for col_num, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            max_length = max(df[col].astype(str).apply(len).max(), len(col)) + 2
            ws.column_dimensions[get_column_letter(col_num)].width = max_length

    return output